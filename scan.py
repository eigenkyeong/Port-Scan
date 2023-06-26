from socket import *
import threading
from openpyxl import load_workbook, Workbook
import ast
import logging

global ip_list, result_list, portList, logger, n, tnum
ip_list = []
result_list = {}
tnum = 200
portList = [1,3,7,9,13,17,19,21,22,23,25,26,37,53,79,80,81,82,88,100,106,110,111,113,119,135,139,143,144,179,199,254,255,280,311,389,427,443,444,445,464,465,497,513,514,515,543,544,548,554,587,593,625,631,636,646,787,808,873,902,990,993,995,1000,1022,1024,1025,1026,1027,1028,1029,1030,1031,1032,1033,1035,1036,1037,1038,1039,1040,1041,1044,1048,1049,1050,1053,1054,1056,1058,1059,1064,1065,1066,1069,1071,1074,1080,1110,1234,1433,1494,1521,1720,1723,1755,1761,1801,1900,1935,1998,2000,2001,2002,2003,2005,2049,2103,2105,2107,2121,2161,2301,2383,2401,2601,2717,2869,2967,3000,3001,3128,3268,3306,3389,3689,3690,3703,3986,4000,4001,4045,4899,5000,5001,5003,5009,5050,5051,5060,5101,5120,5190,5357,5432,5555,5631,5666,5800,5900,5901,6000,6001,6002,6004,6112,6646,6666,7000,7070,7937,7938,8000,8002,8008,8009,8010,8031,8080,8081,8443,8888,9000,9001,9092,9090,9100,9102,9999,10000,10001,10010,32768,32771,49152,49153,49154,49155,49156,49157,50000]

# save log
def save_log():
    global logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(message)s')

    fileHandler = logging.FileHandler('./scanning.log')
    fileHandler.setFormatter(formatter)
    logger.addHandler(fileHandler)

# get ip_list from excel file
def get_ipList():
    global ip_list, result_list, n
    wb = load_workbook("./R53_Domains_public.xlsx")
    ws = wb['Sheet']
    resource = ws["D"]
    ip_list = []
    
    # parsing
    for i in range(1, len(resource)):
        ip = resource[i].value
        if ip.startswith('['):
            ip = ast.literal_eval(ip)
        if type(ip) == list:
            for i in range(len(ip)):
                ip_list.append(ip[i]['Value'])
        else:
            ip_list.append(ip)
    ip_list = set(ip_list)
    n = len(ip_list)
    # print(len(ip_list))

    result_list = dict.fromkeys(ip_list)
    ip_list = list(ip_list)

    for i in range(len(ip_list)):
        result_list[ip_list[i]] = []

# save result to excel file
def save_result():
    global result_list
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = 'result'
    col_names = ['ip', 'port']
    for seq, name in enumerate(col_names):
        ws2.cell(row=1, column=seq+1, value=name)
    
    empty_list = []
    for (ip, port) in result_list.items():
        if len(port) ==  0:
            empty_list.append(ip)

    for ip in empty_list:
        result_list.pop(ip)
        
    row_num = 2
    rn = 0
    for (ip, port) in result_list.items():
        if len(port) == 0:
            continue
        ws2.cell(row=row_num+rn, column=1, value=ip)
        if len(port) == 201:
            ws2.cell(row=row_num+rn, column=2, value='*')
        else:
            for seq, pn in enumerate(port):
                ws2.cell(row=row_num+rn, column=seq+2, value=pn)
        rn = rn + 1

    wb2.save("./result.xlsx")
    wb2.close()


# port scan
class Scan(threading.Thread):
    global portList, ip_list, logger, n, tnum
    def __init__(self, num):
        threading.Thread.__init__(self)
        self.num = num
        if self.num < n % tnum:
            self.round_num = n // tnum
        else:
            self.round_num = n // tnum + 1
    
    def run(self):
        for i in range(self.round_num):
            for port in portList:
                try:
                    idx = self.num + (tnum * i)
                    scanning = socket(AF_INET, SOCK_STREAM)
                    scanning.setsockopt(SOL_SOCKET, SO_REUSEADDR, 1)
                    scanning.connect((ip_list[idx], port))
                    result = scanning.recv(1024)
                    logger.info("result (ip) : " + str(result) + " -> " + ip_list[idx])
                    result_list[ip_list[idx]].append(port)
                    logger.info("ip : " + ip_list[idx] + ", port [" + str(port) + "] open\n")
                    scanning.close()
                except Exception as e:
                    logger.info("[error] ip: " + ip_list[idx] + " , port: " + str(port))
                    # logger.info(e)
                    scanning.close()
                    pass


if __name__=="__main__":

    save_log()
    get_ipList()
    
    # create threads
    threads = []
    for i in range(0, tnum):
        t = Scan(i)
        t.start()
        threads.append(t)

    # thread join
    for thread in threads:
        thread.join()

    logger.info("scanning end\n")
    
    save_result()
    
    logger.info('main exit\n')
    
    